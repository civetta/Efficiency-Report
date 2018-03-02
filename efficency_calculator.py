import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Fill
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import numpy as np
import re


def find_blocks(ws,day):
    max_col = ws.max_column
    max_row = ws.max_row
    col = 3
    create_summary_tables(ws,max_col,day)
    ws.cell(row=1,column=max_col+7,value =day+" Day Summary").fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=14,column=max_col+7,value =day+" Night Summary").fill=PatternFill("solid", fgColor="c6c0ed")
    while col <= max_col:
        teacher_name=ws.cell(row=1,column=col).value
        row = 2
        create_tables(teacher_name,col,max_col,ws,day)
        time_block, tab_block, cell_block = [], [], []
        block_on=False
        block_count=0
        while row < max_row:
            current_cell = ws.cell(row = row,  column = col)
            if current_cell.font.bold==True:
                block_on=True
                time=ws.cell(row = row, column = 1).value
                time=time[8:]
                time_block.append(time)
                tab_block.append(ws.cell(row = row, column = 2).value)
                cell_block.append(current_cell.value)
            if current_cell.font.bold==False:
                if block_on==True:
                    block_count=block_count+1
                    block_tables(col,max_col,time_block,tab_block,cell_block,block_count,ws,teacher_name,day)
                    time_block, tab_block, cell_block = [], [], []
                    block_on=False
            row=row+1  
        col=col+1
                  

def create_tables(teacher_name,col,max_column,ws,day):
    col=col-2
    s_row=(col*8)-7
    ws.cell(row=s_row,column=max_column+2, value = teacher_name)
    ws.cell(row=s_row,column=max_column+3, value = "Average Students")
    ws.cell(row=s_row,column=max_column+4, value ="Average Tabby")
    ws.cell(row=s_row,column=max_column+5, value ="Students/Tabby")
    ws.cell(row=2,column=max_column+7+col, value=teacher_name).fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=15,column=max_column+7+col, value=teacher_name).fill=PatternFill("solid", fgColor="c6c0ed")

def create_summary_tables(ws,max_column,day):

    
    ws.cell(row=2,column=max_column+7, value ="Name").fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=3,column=max_column+7, value =str(day)+" Day Average").fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=4,column=max_column+7, value ="Block 1").fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=5,column=max_column+7, value ="Block 2").fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=6,column=max_column+7, value ="Block 3").fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=7,column=max_column+7, value ="Block 4").fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=8,column=max_column+7, value ="Block 5").fill=PatternFill("solid", fgColor="f7d28a")
    ws.cell(row=9,column=max_column+7, value ="Block 6").fill=PatternFill("solid", fgColor="f7d28a")

    
    ws.cell(row=15,column=max_column+7, value ="Name").fill=PatternFill("solid", fgColor="c6c0ed")
    ws.cell(row=16,column=max_column+7, value =str(day)+" Night Average").fill=PatternFill("solid", fgColor="c6c0ed")
    ws.cell(row=17,column=max_column+7, value ="Block 1").fill=PatternFill("solid", fgColor="c6c0ed")
    ws.cell(row=18,column=max_column+7, value ="Block 2").fill=PatternFill("solid", fgColor="c6c0ed")
    ws.cell(row=19,column=max_column+7, value ="Block 3").fill=PatternFill("solid", fgColor="c6c0ed")
    ws.cell(row=20,column=max_column+7, value ="Block 4").fill=PatternFill("solid", fgColor="c6c0ed")
    ws.cell(row=21,column=max_column+7, value ="Block 5").fill=PatternFill("solid", fgColor="c6c0ed")
    ws.cell(row=22,column=max_column+7, value ="Block 6").fill=PatternFill("solid", fgColor="c6c0ed")

    

def block_tables(col,max_column,time_block,tab_block,cell_block,block_count,ws,teacher_name,day):
    col=col-2
    s_row=(col*8)-7
    ws.cell(row=s_row+block_count,column=max_column+2, value =time_block[0]+" - "+time_block[-1])
    
    ws.cell(row=s_row+block_count,column=max_column+3, value =round(np.average(cell_block),2))
    tab_block=np.array(tab_block).astype(np.float)
    if sum(tab_block)>2:
        check=False
        ws.cell(row=s_row+block_count,column=max_column+4, value =round(sum(tab_block)/len(tab_block),2))
        ws.cell(row=s_row+block_count,column=max_column+5, value =round(round(np.average(cell_block),2)/round(np.average(tab_block),2),2))
        for a in time_block:
            a=str(a)
            a=a.strip()
            day=(a[:a.index(" ")]).strip()
            a=a[a.index(" "):]
            hr=int(a[:a.index(":")])
            m=a[-2:]
            if hr>=8 and hr<12 and m=="PM" or day=="Sat" or day =="Sun":
                ws.cell(row=s_row+block_count,column=max_column+4).fill=PatternFill("solid", fgColor="c6c0ed")
                ws.cell(row=s_row+block_count,column=max_column+5).fill=PatternFill("solid", fgColor="c6c0ed")
                ws.cell(row=s_row+block_count,column=max_column+2).fill=PatternFill("solid", fgColor="c6c0ed")
                ws.cell(row=block_count+16,column=max_column+7+col,value =round(round(np.average(cell_block),2)/round(np.average(tab_block),2),2))
                check=True
                break
        if check==False:
            ws.cell(row=block_count+3,column=max_column+7+col,value =round(round(np.average(cell_block),2)/round(np.average(tab_block),2),2))



def format_sheets(ws,max_column):
    ws.column_dimensions["A"].width = int(25)
    for i in range(2,max_column+1):
        ws.column_dimensions[get_column_letter(i)].width  =  int(15)
    for a in range(max_column+3,max_column+6):
        ws.column_dimensions[get_column_letter(a)].width  =  int(15)
    ws.column_dimensions[get_column_letter(max_column+2)].width  =  int(30)
    ws.freeze_panes = 'A2'

def end_format(ws,max_column):
    new_max=ws.max_column
    empty_col1=get_column_letter(max_column+1)
    empty_col2=get_column_letter(max_column+6)
    for a in range(max_column+6,new_max+1):
        ws.column_dimensions[get_column_letter(a)].width  =  int(15)
    img = Image('arrow.png')
    img2 = Image('arrow.png')
    ws.column_dimensions[empty_col1].width  =  int(20)
    ws.column_dimensions[empty_col2].width  =  int(20)
    ws.add_image(img,empty_col1+"2" )
    ws.add_image(img2,empty_col2+"2" )

    
def create_block_table(wb,lead_name):
    week=["Mon", "Tue", "Wed","Thu","Fri","Sat","Sun"]
    for day in week:
        try:
            ws = wb.get_sheet_by_name(day)
            max_column = ws.max_column
            format_sheets(ws,max_column)
            max_col=find_blocks(ws,day)
                
            
        except Exception as e:
            print e
            continue
        end_format(ws,max_column)
    return max_column
