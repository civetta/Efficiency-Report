from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


def organize_data(ws, start, end, column, block_list, tab_list, max_col, checks, scores):
    """Declares all of the variables needed to create and organize the
    efficiency table"""
    average_student = round(sum(block_list)/float(len(block_list)), 2)
    average_tabby = round(sum(tab_list)/float(len(tab_list)), 2)
    block_escore = round(average_student/float(average_tabby), 2)
    teacher_name = str(ws.cell(row=1, column=column).value)
    start_time = str(ws.cell(row=start, column=6).value)
    end_time = str(ws.cell(row=end, column=6).value)
    time_range = create_time_range(start_time, end_time, ws)
    if '*' in time_range:
        checks['Night Check'] = True
    if '*' not in time_range:
        checks['Day Check'] = True
    paste_list = [time_range, average_student, average_tabby, block_escore]
    paste_data(ws, paste_list, teacher_name, max_col, scores)
    return checks


def create_time_range(start, end, ws):
    """
    Input: Two string of time. 
    Example: 03/21/18 Thu 8:27 AM and 03/21/18 Thu 9:48 AM.  
    Output: A string with just the hours (example: 8:27 AM - 9:48 AM").
    Times is marked with an asteriks if one of the times is after 8 AM
    but before 1AM. Or if the day of the week is Saturday or Sunday
    These mark the cells so that they are calculated differently then
    the "day time" teachers.
    """
    start = datetime.strptime(start, '%m/%d/%y %a %I:%M %p')
    end = datetime.strptime(end, '%m/%d/%y %a %I:%M %p')
    nightshift_start = start.replace(hour=20, minute=0)
    nightshift_end = start.replace(day=start.day+1, hour=2, minute=0)     
    time_range = ""
    if end.weekday() >= 5:
        time_range = '*'
    elif end > nightshift_start and end < nightshift_end:
        time_range = '*'
    return time_range+start.strftime("%I:%M %p")+"-"+end.strftime("%I:%M %p")


def paste_data(ws, paste_list, teacher_name, max_col, scores):
    """Paste the organized data under the daily summary tables.
    It also outlines the cell if it has an asteriks(*)
    which indiciates a night or weekend shift."""  
    night_check = False
    if teacher_name is not None:
        starting_row = find_table(ws, teacher_name, max_col)
        empty_row = find_empty_row(ws, starting_row)
        if '*' in paste_list[0]:
            night_time_teacher(ws, empty_row)
            night_check = True
        if paste_list[-1] is not None:
            color = coniditional_format_row(ws, empty_row, paste_list, scores, night_check)
        for i in range(4):
            current_cell = ws.cell(row=empty_row, column=i+1)
            current_cell.value = paste_list[i]
            current_cell.fill = PatternFill("solid", fgColor=color)


def find_table(ws, teacher_name, max_col):
    """Finds the location of the teacher's daily summary table in the column"""
    for col in range(7, max_col+1):
        if ws.cell(row=1, column=col).value == teacher_name:
            return ((col-7)*8)-6


def find_empty_row(ws, starting_row):
    """Finds the first empty row in the teacher's daily summary table 
    so it can paste the new data in it"""
    for row in range(starting_row, ws.max_row):
        if ws.cell(row=row, column=1).value is None:
            return row


def coniditional_format_row(ws, empty_row, paste_list, scores, night_check):
    """Using the condition_list as scores, it conditionally formats
    "good scores' as green, 'bad scores' as red, and too good scores as
    blue"""
    good_score = scores.get("Good Score")
    upper_bound = scores.get("Upper Bound")
    e_score = paste_list[-1]
    if e_score < good_score:
        return 'f2b8ea'
    elif e_score > upper_bound:
        return 'c0f7f4'
    else:
        return 'dff7c0'


def night_time_teacher(ws, empty_row):
    """Creates thick borders around night time teacher cells"""
    thick_border = Border(left=Side(border_style='thick', color='1F49A1'),
                right=Side(border_style='thick', color='1F49A1'),
                top=Side(border_style='thick', color='1F49A1'),
                bottom=Side(border_style='thick', color='1F49A1'))
    for i in range(4):
        current_cell = ws.cell(row=empty_row, column=i+1)
        current_cell.border = thick_border

