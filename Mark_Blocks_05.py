from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from Calculate_Block_Escores_5A import organize_data
from Create_Tables_04A import create_tables
from Create_Tables_04A import format_sheet


def define_blocks(wb):
    """Goes through each ws and then goes through each column looking for the 
    start and end of each block. It then calls the bolder function
    conditionally format the cells"""
    week = wb.get_sheet_names()
    week = week[:-3]
    for day in week:
        ws = wb.get_sheet_by_name(day)
        create_tables(ws)
        format_sheet(ws)
        max_row, max_col = ws.max_row, ws.max_column
        col = 8
        start_row_to_look = 2
        while col <= max_col:
            start = find_blocks(ws, col, max_row, start_row_to_look, 'start')
            if start >= start_row_to_look and start != max_row:
                end = find_blocks(ws, col, max_row, start, 'end')
                start_row_to_look = end
                bolder(ws, start, end, col, max_col)                  
            else:
                col = col+1
                start_row_to_look = 2
    return wb


def find_blocks(ws, col, max_row, starting_row, position):
    """Looks for either two 0's in a row for the end of a block,or 
    two sequential non zeros for the start of the block"""
    for a in range(starting_row, max_row):
        val1 = ws.cell(row=a, column=col).value
        if position == 'start':  
            if int(val1) > 0:
                val2 = int(ws.cell(row=a+1, column=col).value)
                val3 = int(ws.cell(row=a+2, column=col).value)
                if val2 > 0 or val3 > 0:
                    return a
        elif position == 'end':
            if int(val1) == 0:
                val2 = int(ws.cell(row=a+1, column=col).value)
                val3 = int(ws.cell(row=a+2, column=col).value)
                if val2 == 0 and val3 == 0:
                    return a
    return max_row


def bolder(ws, start, end, column, max_col):
    """This function goes through and bolds and conditionally
    formatts each of the blocks. The bolding will be used later to identify
    a block and for calculating"""            
    tab_list = []
    block_list = []
    for r in range(start, end):
        current_cell = ws.cell(row=r,  column=column)
        current_value = current_cell.value
        Tabby_Cell = int(ws.cell(row=r,  column=7).value)
        plus_1_check = current_value == Tabby_Cell+1 
        minus_1_check = current_value == Tabby_Cell-1
        current_cell.font = Font(bold=True)
        if current_value == Tabby_Cell or plus_1_check or minus_1_check:
            current_cell.fill = PatternFill("solid", fgColor='dff7c0')
        elif current_value < Tabby_Cell and current_value > 0:
            current_cell.fill = PatternFill("solid", fgColor='f2b8ea')
        elif current_value >= Tabby_Cell+2:
            current_cell.fill = PatternFill("solid", fgColor='c0f7f4')
        else:
                continue
        tab_list.append(Tabby_Cell)
        block_list.append(current_value)
    organize_data(ws, start, end, column, block_list, tab_list, max_col)
