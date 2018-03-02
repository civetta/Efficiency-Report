import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Fill
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
import numpy as np
import re

def find_average(ws,max_col,col,start_row,start_col):
    while col <= max_col+1:
        num_list=[]
        row=start_row
        while row<start_row+9:
            o_cell = ws.cell(row=row, column=col)
            if o_cell.value != None:
                try:
                    value = int(o_cell.value)
                    num_list.append(o_cell.value)
                except ValueError:
                    pass
            row=row+1
        if len(num_list)>0:
            ws.cell(row=start_row+2,column=col,value = round(sum(num_list)/len(num_list),2))
            ws.cell(row=start_row+2,column=col).font=Font(bold=True)

            """if start_row<4:
                ws.cell(row=2,column=col-start_col-5,value = round(sum(num_list)/len(num_list),2))
                ws.cell(row=2,column=col-start_col-5).font=Font(bold=True)
            else:
                ws.cell(row=3,column=col-start_col-5,value = round(sum(num_list)/len(num_list),2))
                ws.cell(row=3,column=col-start_col-5).font=Font(bold=True)"""
                

        col=col+1
        
def formatter(ws,start_col):
    max_col=ws.max_column
    ws.column_dimensions['A'].width=30
    for a in range(1,max_col+2):
        ws.column_dimensions[get_column_letter(a)].width  =  int(15)
def create_summary(wb,start_col):
    thin_border = Border(left=Side(border_style='thin',color='afafaf'),
                 right=Side(border_style='thin',color='afafaf'),
                 top=Side(border_style='thin',color='afafaf'),
                 bottom=Side(border_style='thin',color='afafaf'))
    week=["Mon", "Tue", "Wed","Thu","Fri","Sat","Sun"]
    weekly_breakdown = wb.create_sheet("Weekly Breakdown",0)
    
    day_row=1
    for day in week:
        try:
            ws=wb.get_sheet_by_name(day)
            max_col=ws.max_column
            col=start_col+8
            find_average(ws,max_col,col,14,start_col)
            find_average(ws,max_col,col,1,start_col)
            """for b in range(1,start_col+1):
                ws.cell(row=3,column=b).fill=PatternFill("solid", fgColor="c6c0ed")
                ws.cell(row=2,column=b).fill=PatternFill("solid", fgColor="f7d28a")"""
            while col <= max_col+1:
                new_row=day_row
                row=1
                while row<9:
                    o_daycell = ws.cell(row=row, column=col-1)
                    n_daycell = weekly_breakdown.cell(row=new_row,column=col-7-start_col)
                    n_daycell.value=o_daycell.value
                    n_daycell.fill=PatternFill("solid", fgColor="f7d28a")
                    n_daycell.border=thin_border
                    if o_daycell.font==Font(bold=True):
                        n_daycell.font=Font(bold=True)
                    o_nightcell=ws.cell(row=row+13,column=col-1)
                    n_nightcell = weekly_breakdown.cell(row=new_row,column=col-6)
                    n_nightcell.value=o_nightcell.value
                    n_nightcell.fill=PatternFill("solid", fgColor="c6c0ed")
                    n_nightcell.border=thin_border
                    if o_nightcell.font==Font(bold=True):
                        n_nightcell.font=Font(bold=True)
                    row=row+1
                    new_row=new_row+1
                col=col+1
            day_row=new_row+3
            
        except:
            continue
        formatter(weekly_breakdown,start_col)


