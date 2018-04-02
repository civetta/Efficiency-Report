import openpyxl
from openpyxl import Workbook
from datetime import datetime, date
import calendar
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


def create_books(max_col,wb):
    for teacher_col in range(3,max_col):
        current_teacher = Workbook()
        set_up(current_teacher)
        teacher_sheet=current_teacher.create_sheet('Data')
        summary=current_teacher.create_sheet('Weekly Summary')
        find_summary(summary,max_col,teacher_col,wb)
        new_col=1
        for i in range(4,(len((wb.get_sheet_names())))):
            ws = wb.worksheets[i]
            new_col=find_columns(teacher_sheet,ws,new_col,teacher_col)
            
        teacher_name=ws.cell(row=1,column=teacher_col).value
        weekly_sum(summary,teacher_name,wb)
        current_teacher.save(teacher_name+'_.xlsx')
    return current_teacher


def set_up(current_teacher):
    info=current_teacher.create_sheet('FAQ')
    img = Image('faq.png')
    info.add_image(img,'A1' )
    std=current_teacher.get_sheet_by_name('Sheet')
    current_teacher.remove_sheet(std)    
    
    

def weekly_sum(summary,teacher_name,wb):
    ws=wb.get_sheet_by_name('Summary')
    new_col_count=7
    for column in range(1,ws.max_column):
        title=ws.cell(row=2,column=column).value
        if title==teacher_name or title=='Name':
            summary.column_dimensions[get_column_letter(new_col_count)].width  =  int(20)
            for row in range(1,12):
                old_cell=ws.cell(row=row,column=column)
                new_cell=summary.cell(row=row,column=new_col_count)
                copier(old_cell,new_cell)
            new_col_count=new_col_count+1
            

def copier(old_cell,new_cell):
    new_cell.value=old_cell.value
    new_cell.font = copy(old_cell.font)
    new_cell.fill = copy(old_cell.fill)


def find_summary(summary,max_col,teacher_col,wb):
    summary_row_count=1
    for i in range(4,(len((wb.get_sheet_names())))):
            ws = wb.worksheets[i]
            teacher_name=ws.cell(row=1,column=teacher_col).value
            for row in range(1,100):
                try:
                    if teacher_name in ws.cell(row=row,column=max_col+2).value:
                        summary_row_count=copy_summary(row,max_col+2,ws,summary,summary_row_count)
                except:
                    pass


def copy_summary(row,col,ws,summary,summary_row_count):
    for r in range (row,row+8):
        sum_col_count=1
        summary.column_dimensions[get_column_letter(sum_col_count)].width  =  int(30)
        for column in range(col,col+4):
            cell_content=ws.cell(row=r,column=column).value
            summary.cell(row=summary_row_count,column=sum_col_count).value=cell_content
            sum_col_count=sum_col_count+1
            summary.column_dimensions[get_column_letter(sum_col_count)].width  =  int(15)
        summary_row_count=summary_row_count+1
    return summary_row_count
    
def format_columns(teacher_sheet,new_col,ws,teacher_col,row,grayFill):
    
    teacher_sheet.cell(row=2,column=new_col,value=ws.cell(row=1,column=1).value)
    teacher_sheet.column_dimensions[get_column_letter(new_col)].width  =  int(30)
    teacher_sheet.cell(row=2,column=new_col+1,value=ws.cell(row=1,column=2).value)
    teacher_sheet.column_dimensions[get_column_letter(new_col+1)].width  =  int(20)
    teacher_sheet.cell(row=2,column=new_col+2,value=ws.cell(row=1,column=teacher_col).value)
    teacher_sheet.column_dimensions[get_column_letter(new_col+2)].width  =  int(20)
    teacher_sheet.cell(row=2,column=new_col+3).fill=grayFill
    teacher_sheet.cell(row=1,column=new_col+3).fill=grayFill

def copy_columns(teacher_sheet,ws,new_col,teacher_col,sub_count,row):
    teacher_sheet.cell(row=row+2-sub_count,column=new_col,value=ws.cell(row=row,column=1).value)
    teacher_sheet.cell(row=row+2-sub_count,column=new_col+1,value=ws.cell(row=row,column=2).value)
    new_cell=teacher_sheet.cell(row=row+2-sub_count,column=new_col+2)
    old_cell=ws.cell(row=row,column=teacher_col)
    copier(old_cell,new_cell)
    
def create_title(teacher_sheet,new_col):
    title=teacher_sheet.cell(row=1,column=new_col)
    title.font = title.font.copy(bold=True, size=20)
    date=teacher_sheet.cell(row=3,column=new_col).value[:8]
    my_date=datetime.strptime(date,'%m/%d/%y')
    title.value=calendar.day_name[my_date.weekday()]

def find_columns(teacher_sheet,ws,new_col,teacher_col):
    check=False
    num=0
    sub_count=0
    grayFill=PatternFill("solid", fgColor="efefef")
    for row in range(2,75):
        try:
            num=int(ws.cell(row=row,column=teacher_col).value)
        except:
            pass
        if check == False:
            sub_count=sub_count+1
            if num > 0:
                check=True
                format_columns(teacher_sheet,new_col,ws,teacher_col,row,grayFill)
        if check ==True:
            copy_columns(teacher_sheet,ws,new_col,teacher_col,sub_count,row)
            create_title(teacher_sheet,new_col)
            
            teacher_sheet.cell(row=row+2-sub_count,column=new_col+3).fill=grayFill
    if check==True:
        return new_col+4
    else:
        return new_col                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                  
