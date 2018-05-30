import openpyxl
from openpyxl.utils import get_column_letter
"""This module takes the YTD numbers from Raw Pulls sheet and finds the differences every 6 minutes.
It then puts those differences in the Raw Changes worksheet"""

#Copies the first two columns which are the dates and tabby
def first_two_columns(old,new,max_row):
    for row in range(1,max_row):
        for col in range(1,3):
                val=old.cell(row=row,column=col).value
                try:
                    val=int(val)
                except:
                    val=str(val)
                new.cell(row=row,column=col,value=val)

#Copies the first row which is just the names of the teachers
def first_row(old,new):
    for col in range(1,old.max_column+1):
        new.cell(row=1,column=col,value=old.cell(row=1,column=col).value)

#Find the actual difference between the values. Tells how many students a teacher closed every 6 minutes
def find_difference(old,new,max_row, max_col):
    for row in range(3,max_row):
        for col in range(3,max_col):
            cell2=(old.cell(row=row,column=col).value)
            cell1=(old.cell(row=row-1,column=col).value)
            try:
                difference=int(cell2)-int(cell1)
            except:
                difference=None
            new.cell(row=row,column=col,value=difference)

#Calls all of the above functions and formats the columns           
def make_time_difference_sheet(wb):
    old = wb.get_sheet_by_name("Raw Pulls")
    new = wb.create_sheet('Raw Changes')    
    max_row=old.max_row+1
    max_col=old.max_column+1
    first_two_columns(old,new,max_row)
    first_row(old,new)
    find_difference(old,new,max_row,max_col)
    old.column_dimensions[get_column_letter(1)].width  =  int(25)
    new.column_dimensions[get_column_letter(1)].width  =  int(25)
    return wb
