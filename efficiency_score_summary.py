from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
"""The abrivation curr is short for "current". It's used to talk about the 
current column, cell, or row in the ws page."""


def create_summary_page(wb, data_dict, checks):
    """Create the ws page and locates it at the beginning of the workbook.
    Uses the sheet names to make date columns. Then it calls the first sheet
    and uses that to create the names."""
    night_color = 'c0b8f2'
    day_color = 'f2c0b8'
    sheet_list = wb.get_sheet_names()[:-2]
    num_of_days = len(sheet_list)
    wb.create_sheet('Summary', 0)
    ws = wb.get_sheet_by_name('Summary')
    create_title(ws, sheet_list)
    if checks['Day Check'] is True and checks['Night Check'] is True:
        create_table(ws, 3, 'Day ws', data_dict, num_of_days, day_color, wb)
        create_table(ws, 8+num_of_days, 'Night ws', data_dict, num_of_days, night_color, wb)
    elif checks['Day Check'] is True and checks['Night Check'] is False:
        create_table(ws, 3, 'Day ws', data_dict, num_of_days, day_color, wb)
    else:
        create_table(ws, 3, 'Night ws', data_dict, num_of_days, night_color, wb)


def create_title(ws, sheet_list):
    """Creates and formats the title of the ws page. Uses 
    worksheet dates to create title"""
    date_range = sheet_list[0]+" - "+sheet_list[-1]
    title_of_ws = "Summary Of Days: " + date_range
    ws.cell(row=1, column=1).value = title_of_ws
    ws.cell(row=1, column=1).font = Font(size=30, bold=True)
    ws.row_dimensions[1].height = int(60)


def create_table(ws, header_row, table_name, data_dict, num_of_days, color, wb):
    """Uses the names from the first non ws page and copies and pastes
    them into the ws page, using a bit of formatting"""
    create_sub_titles(ws, header_row, table_name, num_of_days, color)
    create_date_column(wb, ws, color, header_row, num_of_days) 
    create_teacher_header_row(ws, header_row, wb)
    for col in range(2, ws.max_column):
        column_array = []
        for row in range(header_row+1, header_row+num_of_days+1):
            curr_cell = ws.cell(row=row, column=col)
            format_curr_cell(curr_cell, color)
            data = paste_data(
                curr_cell, ws, col, row, header_row, data_dict, table_name)
            if data != '':
                column_array.append(data)
        paste_average(column_array, header_row, num_of_days, col, ws)
    ws.column_dimensions['A'].width = int(20)


def create_sub_titles(ws, header_row, table_name, num_of_days, color):
    """Creates the sub titles, like "Day ws" or "night ws"."""
    ws.row_dimensions[header_row].height = int(30)
    title_of_table = ws.cell(row=header_row, column=1)
    format_curr_cell(title_of_table, color)
    total_average_title = ws.cell(row=header_row+num_of_days+1, column=1)
    big_font(title_of_table, table_name)
    big_font(total_average_title, 'Total Average')
    ws.row_dimensions[header_row].height = int(40)


def create_teacher_header_row(ws, header_row, wb):
    """Creates the header row for each table. The header row consistants
    of all teacher names with a new line in between their first and 
    laste name"""
    rawsheet = wb.get_sheet_by_name('Raw Pulls')
    for col_in_rawsheet in range(3, rawsheet.max_column):
        curr_col = col_in_rawsheet-1
        teacher_name = rawsheet.cell(row=1, column=col_in_rawsheet).value
        first_name = teacher_name[:teacher_name.index(" ")]
        last_name = teacher_name[teacher_name.index(" ")+1:]
        teacher_name_formatted = first_name+'\r\n'+last_name
        current_cell = ws.cell(row=header_row, column=curr_col)
        big_font(current_cell, teacher_name_formatted)
        ws.column_dimensions[get_column_letter(curr_col)].width = int(20)


def create_date_column(wb, ws, color, header_row, num_of_days):
    """Creates the date column, which represents the y axis of the ws
    table. Each row in the column is a date. It uses the list of sheet 
    names to create this"""
    dates = wb.get_sheet_names()[1:-2]
    count = 0
    for row in range(header_row+1, header_row+num_of_days+1):
        date_cell = ws.cell(row=row, column=1)
        date_cell.value = dates[count]
        format_curr_cell(date_cell, color)
        count = count+1 


def paste_data(curr_cell, ws, column, row, header_row, data_dict, table_name):
    """Using the current row for the date, and column to find teacher name
    it looks into the data_dict and matches."""
    date = ws.cell(row=row, column=1).value
    teacher_name = ws.cell(row=header_row, column=column).value
    teacher_name = teacher_name.replace('\r\n', ' ')
    date_dictionary = data_dict[date]
    teacher_data = date_dictionary[teacher_name]
    if "Day" in table_name:
        data = teacher_data['Day Average']
        curr_cell.value = data
    else:
        data = teacher_data['Night Average']
        curr_cell.value = data
    return data
    

def paste_average(column_array, header_row, num_of_days, col, ws):
    if len(column_array) > 0:
        average = round(sum(column_array)/len(column_array), 2)
        avg_cell_row = header_row+num_of_days+1
        average_cell = ws.cell(row=avg_cell_row, column=col)
        big_font(average_cell, average)


def format_curr_cell(curr_cell, color):
    thin_border = Border(left=Side(border_style='thin', color='E6E6E6'),
        right=Side(border_style='thin', color='E6E6E6'),
        top=Side(border_style='thin', color='E6E6E6'),
        bottom=Side(border_style='thin', color='E6E6E6'))
    curr_cell.fill = PatternFill("solid", fgColor=color)
    curr_cell.border = thin_border


def big_font(cell, value):
    cell.value = value
    cell.font = Font(size=15, bold=True)
    cell.alignment = Alignment(wrapText=True)
