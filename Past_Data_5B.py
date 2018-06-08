from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
condition_list = {"Good Score": float(.90), "Upper Bound": float(1.25)}


def paste_data(ws, paste_list, teacher_name, max_col):
    """Paste the organized data under the daily summary tables.
    It also outlines the cell if it has an asteriks(*)
    which indiciates a night or weekend shift."""  

    if teacher_name is not None:
        starting_row = find_table(ws, teacher_name, max_col)
        empty_row = find_empty_row(ws, starting_row)
        if paste_list[-1] is not None:
            color_to_fill = coniditional_format_row(ws, empty_row, paste_list)
        if '*' in paste_list[0]:
            night_time_teacher(ws, empty_row)
        for i in range(4):
            current_cell = ws.cell(row=empty_row, column=i+1)
            current_cell.value = paste_list[i]
            current_cell.fill = PatternFill("solid", fgColor=color_to_fill)


def find_table(ws, teacher_name, max_col):
    """Finds the location of the teacher's daily summary table in the column"""
    for col in range(7, max_col+1):
        if ws.cell(row=1, column=col).value == teacher_name:
            return ((col-7)*8)-6


def find_empty_row(ws, starting_row):
    """Finds the first empty row in the teacher's daily summary table 
    so it can paste the new data in it"""
    for row in range(starting_row, ws.max_row):
        if ws.cell(row=row, column=1).value is None:
            return row


def coniditional_format_row(ws, empty_row, paste_list):
    """Using the condition_list as scores, it conditionally formats
    "good scores' as green, 'bad scores' as red, and too good scores as
    blue"""
    good_score = condition_list.get("Good Score")
    upper_bound = condition_list.get("Upper Bound")
    e_score = paste_list[-1]
    if e_score < good_score:
        return 'f2b8ea'
    elif e_score > upper_bound:
        return 'c0f7f4'
    else:
        return 'dff7c0'


def night_time_teacher(ws, empty_row):
    """Creates thick borders around night time teacher cells"""
    thick_border = Border(left=Side(border_style='thick', color='1F49A1'),
                right=Side(border_style='thick', color='1F49A1'),
                top=Side(border_style='thick', color='1F49A1'),
                bottom=Side(border_style='thick', color='1F49A1'))

    for i in range(4):
        current_cell = ws.cell(row=empty_row, column=i+1)
        current_cell.border = thick_border
