from openpyxl import Workbook
from openpyxl.drawing.image import Image
from copy import copy
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def create_books(wb):
    """Collects all of the worksheet variables that will be used and creates
    a new workbook which is defined as just the teacher name. Then it creates
    an FAQ page, a Data page, and a Summary page."""
    sheet_list = wb.get_sheet_names()
    summaryws = wb.get_sheet_by_name(sheet_list[0])
    """Goes through all of the teacher names listed in summary worksheet, with
    the understanding the first table will always have it's teacher header
    row at row 3."""
    for teacher in range(2, summaryws.max_column+1):
        teacher_name = summaryws.cell(row=3, column=teacher).value
        teacher_name = teacher_name.replace('\r\n', " ")
        teacherbook = Workbook()
        set_up(teacherbook)
        copy_summary(teacherbook, wb, teacher_name)
        copy_data(teacherbook, wb, teacher_name)
        teacherbook.save(teacher_name+'.xlsx')


def set_up(current_teacher):
    """Creates an FAQ page and includes the FAQ image. Then it deletes
    the default sheet that is created when a workbook is created."""
    info = current_teacher.create_sheet('FAQ')
    img = Image('faq.png')
    info.add_image(img, 'A1')
    std = current_teacher.get_sheet_by_name('Sheet')
    current_teacher.remove_sheet(std)  


def copy_summary(teacherbook, wb, teachername):
    teacher_summary = teacherbook.create_sheet('Weekly Summary')

    
def copy_data(teacherbook, wb, teachername):
    """Copies entire columns from the daily worksheets including conditional
    formatting"""
    teacher_data = teacherbook.create_sheet('Data')
    data_column = 1
    days_in_sheet = wb.get_sheet_names()[1:-2]
    for day in days_in_sheet:
        ws = wb.get_sheet_by_name(day)
        column = find_teacher_column(ws, teachername)
        for row in range(1, ws.max_row+1):
            # Copies time column
            old_time = ws.cell(row=row, column=6)
            new_time = teacher_data.cell(row=row+1, column=data_column)
            copier(old_time, new_time)
            # Copies tabby column
            old_tabby = ws.cell(row=row, column=7)
            new_tabby = teacher_data.cell(row=row+1, column=data_column+1)
            copier(old_tabby, new_tabby)
            # Copies actual data column
            old_data = ws.cell(row=row, column=column)
            new_data = teacher_data.cell(row=row+1, column=data_column+2)
            copier(old_data, new_data)
            # Grays out the fourth column which is used as a divide
            divider_col = teacher_data.cell(row=row, column=data_column+3)
            divider_col.fill = PatternFill("solid", fgColor='F2F2F2')
        data_column = data_column+4
    format_data_sheet(teacher_data)


def format_data_sheet(teacher_data):
    """Sets timestamp column to 30 pixes wide, tabby column to 10, the actual
    data column to 20, and the divider column to 20 as well.
    Then it uses the date to create a title for each section."""
    x = 1
    teacher_data.row_dimensions[1].height = int(30)
    while x < teacher_data.max_column+1:
        teacher_data.column_dimensions[get_column_letter(x)].width = int(30)
        teacher_data.column_dimensions[get_column_letter(x+1)].width = int(10)
        teacher_data.column_dimensions[get_column_letter(x+2)].width = int(20)
        teacher_data.column_dimensions[get_column_letter(x+3)].width = int(20)
        date = teacher_data.cell(row=2, column=x).value
        date = date[date.index(" "):]
        teacher_data.cell(row=1, column=x).value = date
        teacher_data.cell(row=1, column=x).font = Font(size=30, bold=True)
        x = x+4

        


def find_teacher_column(ws, teachername):
    """Finds the column in which the 
    teacher's data is located in each worksheet"""
    for column in range(8, ws.max_column+1):
        cell = ws.cell(row=1, column=column).value
        if cell == teachername:
            return column


def copier(old_cell, new_cell):
    new_cell.value = old_cell.value
    new_cell.font = copy(old_cell.font)
    new_cell.fill = copy(old_cell.fill)