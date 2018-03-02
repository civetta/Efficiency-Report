import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Fill
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

import numpy as np
import re

def find_row(wb):
    breakdown = wb.get_sheet_by_name("Weekly Breakdown")
    max_row=breakdown.max_row
    summary=create_summary(wb,breakdown)
    current_row=1
    past_row=3
    while current_row<max_row:
        current_cell=breakdown.cell(row=current_row, column=1).value
        if current_cell != None:
            if "Day Average" in current_cell:
                copy_row(past_row,breakdown,summary,current_row)
                summary.cell(row=past_row, column=1, value=breakdown.cell(row=current_row-2, column=1).value)
                past_row=past_row+1
        current_row=current_row+1
    find_average(summary)
    formatter(summary)

def formatter(summary):
    thin_border = Border(left=Side(border_style='thin',color='afafaf'),
                 right=Side(border_style='thin',color='afafaf'),
                 top=Side(border_style='thin',color='afafaf'),
                 bottom=Side(border_style='thin',color='afafaf'))
    max_col=summary.max_column
    max_row=summary.max_row
    for a in range(1,max_col+1):
        summary.column_dimensions[get_column_letter(a)].width  =  int(17)
        summary.cell(row=max_row,column=a).font=Font(size=15,bold=True)
        
        for row in range(1,max_row+1):
            summary.cell(row=row,column=a).border=thin_border
            summary.cell(row=row,column=a).alignment = Alignment(wrapText=True)
    summary.cell(row=1,column=1).font=Font(size=18,bold=True)
    summary.column_dimensions[get_column_letter(1)].width  =  int(25)
                
                
        
    

def copy_row(past_row,breakdown,summary,row):
    i=1
    while i <= breakdown.max_column:
        summary.cell(row=past_row, column=i, value=breakdown.cell(row=row, column=i).value)
        if breakdown.cell(row=row,column=i).fill==PatternFill("solid", fgColor="f7d28a"):
            summary.cell(row=past_row, column=i).fill=PatternFill("solid", fgColor="f7d28a")
        if breakdown.cell(row=row,column=i).fill==PatternFill("solid", fgColor="c6c0ed"):
            summary.cell(row=past_row, column=i).fill=PatternFill("solid", fgColor="c6c0ed")
        i=i+1

def create_summary(wb,breakdown):
    summary = wb.create_sheet("Summary", 0)
    summary.cell(row=10, column=1, value="Weekly Average")
    raw_sheet = wb.get_sheet_by_name("Raw Changes")
    first_date=str(raw_sheet.cell(row=2,column=1).value)
    last_date=str(raw_sheet.cell(row=raw_sheet.max_row-3,column=1).value)
    summary.cell(row=1,column=1,value="Weekly Summary")
    summary.cell(row=1, column=2, value=first_date[:first_date.index(" ")]+" - "+last_date[:last_date.index(" ")])
    
    i=1
    while i <=breakdown.max_column:
        current_cell = breakdown.cell(row=2, column=i).value
        summary.cell(row=2, column=i,value=breakdown.cell(row=2, column=i).value)
        i=i+1
    return summary
                
def find_average(summary):
    max_col=summary.max_column
    col=2
    while col<=max_col:
        average_list=[]
        row=3
        while row <= 9:
            celler=summary.cell(row=row,column=col).value
            if celler != None:
                try:
                    value = int(celler)
                    average_list.append(celler)
                except ValueError:
                    pass
            row=row+1
        if len(average_list)>1:
            av=(sum(average_list)/len(average_list))
            summary.cell(row=10,column=col, value=round(av,2))
        if len(average_list)is 1:
            summary.cell(row=10,column=col, value=average_list[0])
        col=col+1
            

