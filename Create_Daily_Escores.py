from Past_Data_5B import find_empty_row
from openpyxl.styles import Font
#from efficiency_score_summary import paste_into_summary_table


def find_non_empty_tables(wb):
    """Goes through each worksheet and finds each teacher table and
     the first empty row in each table"""
    week = wb.get_sheet_names()
    week = week[:-2]   
    all_data = {} 
    for day in week:
        day_dict = {}
        ws = wb.get_sheet_by_name(day)
        for col in range(8, ws.max_column+1):
            table_row = ((col-7)*8)-6
            teacher_name = ws.cell(row=1, column=col).value
            empty_row = find_empty_row(ws, table_row)
            if table_row-empty_row != 0:
                create_arrays(ws, table_row, empty_row, teacher_name, day_dict)
                #paste_into_summary_table(teacher_score, day, teacher_name)
        all_data.update({day:day_dict})
    return all_data

def create_arrays(ws, table_row, empty_row, teacher_name, day_dict):
    """Goes through each tabble and add each block efficiency score to 
    either a day array or a night array, using * as an indicator as a
    night shift. It also removes the *"""
    day_array = []
    night_array = []
    for row in range(table_row+1, empty_row):
        block_escore = ws.cell(row=row, column=4).value
        time_range = ws.cell(row=row, column=1).value
        if '*' in time_range:
            night_array.append(block_escore)
            ws.cell(row=row, column=1).value = time_range[1:]
        else:
            day_array.append(block_escore)
    teacher_score = past_daily_escore(ws, empty_row, day_array, night_array, teacher_name, day_dict)
    return teacher_score


def past_daily_escore(ws, empty_row, day_array, night_array, teacher_name, day_dict):
    """Using the array and the empty row (the first empty row in the table),
    it calculates the daily average, pasting it as either a "Day Average" 
    or a "Night Average" or both."""
    if len(day_array) > 0:
        day_average = round(sum(day_array)/len(day_array), 2)
        ws.cell(row=empty_row, column=1, value="Day Average").font = Font(bold=True)
        ws.cell(row=empty_row, column=4, value=day_average).font = Font(bold=True)
        empty_row = empty_row+1
    else:
        day_average = ""
    if len(night_array) > 0:
        night_average = round(sum(night_array)/len(night_array), 2)
        ws.cell(row=empty_row, column=1, value="Night Average").font = Font(bold=True)
        ws.cell(row=empty_row, column=4, value=night_average).font = Font(bold=True)
    else:
        night_average = ""
    return day_dict.update({teacher_name:{'Day Average': day_average, "Night Average": night_average,}})