from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from copy import copy
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side


def copier(old_cell, new_cell):
    '''Copies both the value, font style, and fill from old cell
    into the new cell'''
    new_cell.value = old_cell.value
    if old_cell.has_style:
        new_cell.font = copy(old_cell.font)
        new_cell.fill = copy(old_cell.fill)
        new_cell.alignment = copy(old_cell.alignment)
        new_cell.border = copy(old_cell.border)


def find_teacher_column(ws, teachername, row_to_look, start_column):
    """Finds the column in which the 
    teacher's data is located in each worksheet. It removes 
    new lines if there is one, so that the generic teachername 
    matches the teachername in the summary page."""
    for column in range(start_column, ws.max_column+1):
        cell = ws.cell(row=row_to_look, column=column).value
        if '\r\n' in cell:
            cell = cell.replace('\r\n', " ")
        if cell == teachername:
            return column