from openpyxl import Workbook
from openpyxl.drawing.image import Image
from copy_teacher_data import copy_data
from copy_teacher_summary import copy_summary
from datetime import datetime, timedelta
import os 
import warnings
warnings.filterwarnings("ignore")
def create_books(wb,lead_name,save_date,debug):
    """Collects all of the worksheet variables that will be used and creates
    a new workbook which is defined as teacherbook. Then it creates
    an FAQ page, a Data page, and a Summary page."""
    sheet_list = wb.sheetnames
    summaryws = wb[sheet_list[0]]
    """Goes through all of the teacher names listed in summary worksheet, with
    the understanding the first table will always have it's teacher header
    row at row 3."""
    for teacher in range(2, summaryws.max_column+1):
        teacher_name = summaryws.cell(row=3, column=teacher).value
        teacher_name = teacher_name.replace('\r\n', " ")
        teacherbook = Workbook()
        create_faq(teacherbook)
        copy_summary(teacherbook, wb, teacher_name)
        copy_data(teacherbook, wb, teacher_name)
        
        if debug is False:
                save_teacherbook(teacherbook,teacher_name,lead_name,save_date)
        else:
                teacherbook.save('Output/Teacher Books/'+teacher_name+'.xlsx')

def save_teacherbook(wb,teacher_name,lead_name,save_date):
    path = 'C:\Users\kelly.richardson\OneDrive - Imagine Learning Inc\Reports\Efficiency Reports'
    file_name = teacher_name+"_"+save_date+"-EReport.xlsx"
    save_location = os.path.join(path,lead_name,teacher_name+" E-Report")
    if not os.path.isdir(save_location):
        os.makedirs (save_location)
    save_name = save_location+"/"+file_name
    wb.save(save_name)


def create_faq(current_teacher):
    """Creates an FAQ page and includes the FAQ image. Then it deletes
    the default sheet that is created when a workbook is created."""
    info = current_teacher.create_sheet('FAQ')
    img = Image('faq.png')
    info.add_image(img, 'A1')
    del current_teacher['Sheet']
    


