from daily_ws.format_block_escore import find_table, find_empty_row
from utility import copier, find_teacher_column
from summary_ws.create_summary_tables import big_font
from openpyxl.utils import get_column_letter


def copy_summary(teacherbook, wb, teachername):
    """Goes through each day of the week worksheet and using the 
    "find_table function that is used earlier when calculating daily
    blocks. The find table function returns the row in which teachername
    daily summary table starts. Then it copies the first five columns and 8 rows
    since when creating the tables, they were all 8 rows tall."""
    teach_summary = teacherbook.create_sheet('Weekly Summary')
    copy_daily_tables(teach_summary, wb, teachername)
    copy_summary_page(teach_summary, wb, teachername)
    format_summary_page(teach_summary)


def copy_daily_tables(teach_summary, wb, teachername):
    """Goes through each daily sheet, and copies the teacher summary daily
    tables. If first finds the table using the function used earlier in 
    format block escore (find table), then it copies 8 rows since each
    table is 8 rows tall."""
    days_in_sheet = wb.get_sheet_names()[1:-2]
    teacherbook_row = 5
    for day in days_in_sheet:
        ws = wb.get_sheet_by_name(day)
        max_col = ws.max_column
        start_of_table = find_table(ws, teachername, max_col)
        date = ws.cell(row=2, column=6).value
        date = date[:date.index(" ")]
        date_cell = teach_summary.cell(row=teacherbook_row-1, column=1) 
        big_font(date_cell, date)
        for row in range(start_of_table, start_of_table+8):
            for col in range(1, 5):
                old_cell = ws.cell(row=row, column=col)
                new_cell = teach_summary.cell(row=teacherbook_row, column=col)
                copier(old_cell, new_cell)
            teacherbook_row = teacherbook_row+1


def copy_summary_page(teach_summary, wb, teachername):
    """Copies the teacher column in the Summary ws in Leadbook. It also puts
    the title that is found in Summary ws and make it the title of this 
    summary page."""
    old_summary = wb.get_sheet_by_name('Summary')
    teacher_col = find_teacher_column(old_summary, teachername, 3, 1)
    title = old_summary.cell(row=1, column=1)
    new_title = teach_summary.cell(row=1, column=1)
    copier(title, new_title)
    for row in range(3, old_summary.max_row+1):
        old_date = old_summary.cell(row=row, column=1)
        new_date = teach_summary.cell(row=row, column=7)
        copier(old_date, new_date)
        old_cell = old_summary.cell(row=row, column=teacher_col)
        new_cell = teach_summary.cell(row=row, column=8)
        copier(old_cell, new_cell)


def format_summary_page(teach_summary):
    """Makes all of the columns have a width of 20 and then goes back
    and make the first column which includes the time ranges, 30 pixels
    wide"""
    for a in range(1, teach_summary.max_column+1):
        teach_summary.column_dimensions[get_column_letter(a)].width = int(20)
    teach_summary.column_dimensions['A'].width = int(30)

        

