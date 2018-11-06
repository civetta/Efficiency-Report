from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from utility import copier, find_teacher_column


def copy_data(teacherbook, wb, teachername):
    """Copies entire columns from the daily worksheets including conditional
    formatting"""
    teacher_data = teacherbook.create_sheet('Data')
    data_column = 1
    days_in_sheet = wb.get_sheet_names()[1:-1]
    for day in days_in_sheet:
        start_of_teacher_day = False
        ws = wb.get_sheet_by_name(day)
        column = find_teacher_column(ws, teachername, 1, 8)
        teacher_worked = did_work(ws, column)
        if teacher_worked == True:
            for row in range(1, ws.max_row+1):
                if start_of_teacher_day == False:
                    if ws.cell(row=row, column=column).font.bold == True:
                        start_of_teacher_day = True
                        new_row=3
                if start_of_teacher_day == True or row == 1:
                    # Copies time column
                    if row==1:
                        new_row=2
                    old_time = ws.cell(row=row, column=6)
                    new_time = teacher_data.cell(row=new_row, column=data_column)
                    copier(old_time, new_time)
                    # Copies tabby column
                    old_tabby = ws.cell(row=row, column=7)
                    new_tabby = teacher_data.cell(row=new_row, column=data_column+1)
                    copier(old_tabby, new_tabby)
                    # Copies actual data column

                    old_data = ws.cell(row=row, column=column)
                    new_data = teacher_data.cell(row=new_row, column=data_column+2)
                    copier(old_data, new_data)
                    # Grays out the fourth column which is used as a divide
                    divider_col = teacher_data.cell(row=new_row, column=data_column+3)
                    divider_col.fill = PatternFill("solid", fgColor='F2F2F2')
                    new_row=new_row+1
            date = teacher_data.cell(row=3, column=data_column).value
            date = date[:date.index(" ")]
            teacher_data.cell(row=1, column=data_column).value = date
            teacher_data.cell(row=1, column=data_column).font = Font(size=30, bold=True)        
            teacherbook.save('Output/Teacher Books/'+teachername+'.xlsx')
            data_column = data_column+4
    format_data_sheet(teacher_data)

def did_work(ws, column):
    for row in range(1, ws.max_row+1):
        if ws.cell(row=row, column=column).font.bold == True:
            return True
    return False
        
def format_data_sheet(teacher_data):
    """Sets timestamp column to 30 pixes wide, tabby column to 10, the actual
    data column to 20, and the divider column to 20 as well.
    Then it uses the date to create a title for each section."""
    x = 1
    teacher_data.row_dimensions[1].height = int(30)
    while x < teacher_data.max_column+1:
        teacher_data.column_dimensions[get_column_letter(x)].width = int(30)
        teacher_data.column_dimensions[get_column_letter(x+1)].width = int(10)
        teacher_data.column_dimensions[get_column_letter(x+2)].width = int(20)
        teacher_data.column_dimensions[get_column_letter(x+3)].width = int(20)
        x=x+4


