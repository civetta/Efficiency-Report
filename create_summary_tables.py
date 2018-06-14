from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment


def big_font(cell, value):
    """Small function that takes a a cell, pastes the value into the cell, and
    then set the font to bold, size to 15, and include text wraping"""
    cell.value = value
    cell.font = Font(size=15, bold=True)
    cell.alignment = Alignment(wrapText=True)


def create_sub_titles(ws, header_row, table_name, num_of_days):
    """Creates the sub titles, like "Day Summary" or "night ws"."""
    ws.row_dimensions[header_row].height = int(30)
    title_of_table = ws.cell(row=header_row, column=1)
    total_average_title = ws.cell(row=header_row+num_of_days+1, column=1)
    big_font(title_of_table, table_name)
    big_font(total_average_title, 'Total Average')
    ws.row_dimensions[header_row].height = int(40)


def create_teacher_header_row(ws, header_row, wb):
    """Creates the header row for each table. The header row consistants
    of all teacher names with a new line in between their first and 
    laste name"""
    rawsheet = wb.get_sheet_by_name('Raw Pulls')
    for col_in_rawsheet in range(3, rawsheet.max_column+1):
        curr_col = col_in_rawsheet-1
        teacher_name = rawsheet.cell(row=1, column=col_in_rawsheet).value
        first_name = teacher_name[:teacher_name.index(" ")]
        last_name = teacher_name[teacher_name.index(" ")+1:]
        teacher_name_formatted = first_name+'\r\n'+last_name
        current_cell = ws.cell(row=header_row, column=curr_col)
        big_font(current_cell, teacher_name_formatted)
        ws.column_dimensions[get_column_letter(curr_col)].width = int(20)


def create_date_column(wb, ws, color, header_row, num_of_days):
    """Creates the date column, which represents the y axis of the ws
    table. Each row in the column is a date. It uses the list of sheet 
    names to create this"""
    dates = wb.get_sheet_names()[1:-2]
    count = 0
    for row in range(header_row+1, header_row+num_of_days+1):
        date_cell = ws.cell(row=row, column=1)
        date_cell.value = dates[count]
        count = count+1 


def format_table(ws, header_row, color, num_of_days):
    """Goes through the table and colors it in, with the color"""
    for col in range(1, ws.max_column+1):
        for row in range(header_row, header_row+num_of_days+1):
            curr_cell = ws.cell(row=row, column=col)
            thin_border = Border(left=Side(border_style='thin', color='E6E6E6'),
                right=Side(border_style='thin', color='E6E6E6'),
                top=Side(border_style='thin', color='E6E6E6'),
                bottom=Side(border_style='thin', color='E6E6E6'))
            curr_cell.fill = PatternFill("solid", fgColor=color)
            curr_cell.border = thin_border
