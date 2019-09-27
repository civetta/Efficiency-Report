import pandas as pd 
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
    
def create_dataframe(data_dict, wb,save_date,debug):
    #Data_dict is a dictionary of dicrionarys. It's {date:{teacher{night:escore, day:escore}}}
    df = pd.DataFrame({'Teacher':[], 'Team':[], 'Date':[], 'E-Score':[], 'Month_Num':[], 'Weekday':[], 'Weekday_Num':[]})
    for date, value in data_dict.items():
        date_dict = data_dict[date]
        ws = wb[date]
        #Gets date from Worksheet so it includes the year versus just the ws title which does not.
        full_date = ws.cell(row=2, column=6).value
        #Creates a datetime object and then paress it a bunch of different ways to get values
        dt_obj = datetime.datetime.strptime(full_date, '%m/%d/%y %a %I:%M %p')
        dater = dt_obj.strftime('%m/%d/%y')
        weekday = dt_obj.strftime('%m/%d/%y %A')
        month_num = dt_obj.strftime('%m')
        weekday_num = dt_obj.strftime('%w')
        for teacher, data in date_dict.items():
            team = find_team(teacher)

            
            teacher_dict = date_dict[teacher]
            if "  " in teacher:
                teacher = teacher.replace("  "," ")
            df = df.append(({'Teacher':teacher, 'Team':team, 'Date':dater, 
                'E-Score':(teacher_dict['Night Average']),'Day/Night':'Night', 'Month_Num':month_num, 
                'Weekday':weekday, 'Weekday_Num':weekday_num}), ignore_index=True)
            df = df.append(({'Teacher':teacher, 'Team':team, 'Date':dater, 
                'E-Score':(teacher_dict['Day Average']),'Day/Night':'Day', 'Month_Num':month_num, 
                'Weekday':weekday, 'Weekday_Num':weekday_num}), ignore_index=True)

    df = df[['Teacher', 'Team', 'Date', 'E-Score', 'Day/Night', 'Month_Num', 'Weekday', 'Weekday_Num']]
    create_workbook(df,save_date,debug)

def create_workbook(df,save_date,debug):
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    if debug is True:
        wb.save('Output/Fall/'+save_date+'_Management.xlsx')
    else:
        path = "C:\\Users\kelly.richardson\OneDrive - Imagine Learning Inc\Reports\Efficiency Reports\Teaching Department"
        file_name = save_date+'_Management.csv'
        save_name = os.path.join(path,file_name)
        wb.save(save_name)


def find_team(teacher):
    team_org = { 'Laura Gardiner':[ 'Laura Gardiner', 'Caren Glowa', 'Crystal Boris', 'Jamie Weston', 'Kay Plinta-Howard', 'Marcella Parks', 'Melissa Mitchell', 'Michelle Amigh', 'Stacy Good'],
    'Rachel Adams':['Rachel Adams', 'Clifton Dukes', 'Heather Chilleo', 'Hester Southerland', 'Kelly Richardson', 'Kimberly Stanek', 'Michele  Irwin', 'Nancy Polhemus', 'Juventino Mireles'],
    'Melissa Cox':[ 'Melissa Cox','Andre Lawe', 'Emily McKibben', 'Erica De Coste', 'Erin Hrncir', 'Erin Spilker', 'Jennifer Talaski', 'Julie Horner', 'Lisa Duran', 'Preston Tirey'],
    'Sara  Watkins':['Sara  Watkins', 'Alisa Lynch', 'Andrea Burkholder', 'Angel Miller', 'Bill Hubert', 'Donita Farmer', 'Jessica Connole', 'Laura Craig', 'Nicole Marsula', 'Rachel Romano', 'Veronica Alvarez', 'Wendy Bowser'],
    'Kristin Donnelly':['Kristin Donnelly', 'Carol Kish', 'Erica Basilone', 'Euna Pineda', 'Hannah Beus', 'Jenni Alexander', 'Jessica Throolin', 'Natasha Andorful', 'Nicole Knisely', 'Shannon Stout'],
    'Gabriela Torres':['Gabriela Torres', 'Amy Stayduhar', 'Audrey Rogers', 'Cheri Shively', 'Kathryn Montano', 'Karen Henderson', 'Lynae Shepp', 'Meaghan Wright', 'Veraunica Wyatt']}
    for team_lead, teams in team_org.items():
        if teacher in teams:
            names = team_lead.split(" ")
            return names[0]
