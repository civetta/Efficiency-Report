from openpyxl.styles import Font
from openpyxl.styles import Alignment
from create_summary_tables import create_sub_titles, create_date_column
from create_summary_tables import create_teacher_header_row, format_table
from create_summary_tables import big_font


def create_summary_page(wb, data_dict, checks):
    """Create the ws page and locates it at the beginning of the workbook.
    Uses the sheet names to make date columns. Then it calls the first sheet
    and uses that to create the names."""
    night_color = 'c0b8f2'
    day_color = 'f2c0b8'
    sheet_list = wb.get_sheet_names()[:-1]
    days_num = len(sheet_list)
    wb.create_sheet('Summary', 0)
    ws = wb.get_sheet_by_name('Summary')
    create_title(ws, sheet_list)
    if checks['Day Check'] is True and checks['Night Check'] is True:
        create_table(ws, 3, 'Day Summary', data_dict, days_num, day_color, wb)
        create_table(
            ws, 8+days_num, 'Night Summary', data_dict, days_num, night_color, wb)
    elif checks['Day Check'] is True and checks['Night Check'] is False:
        create_table(ws, 3, 'Day Summary', data_dict, days_num, day_color, wb)
    else:
        create_table(
            ws, 3, 'Night Summary', data_dict, days_num, night_color, wb)
    print data_dict


def create_title(ws, sheet_list):
    """Creates and formats the title of the ws page. Uses 
    worksheet dates to create title"""
    date_range = sheet_list[0]+" - "+sheet_list[-1]
    title_of_ws = "Summary Of Days: " + date_range
    ws.cell(row=1, column=1).value = title_of_ws
    ws.cell(row=1, column=1).font = Font(size=30, bold=True)
    ws.row_dimensions[1].height = int(60)


def create_table(ws, header_row, table_name, data_dict, days_num, color, wb):
    """Uses the names from the first non ws page and copies and pastes
    them into the ws page, using a bit of formatting"""
    create_sub_titles(ws, header_row, table_name, days_num)
    create_date_column(wb, ws, color, header_row, days_num) 
    create_teacher_header_row(ws, header_row, wb)
    format_table(ws, header_row, color, days_num)
    for col in range(2, ws.max_column+1):
        column_array = []
        for row in range(header_row+1, header_row+days_num+1):
            data = paste_data(ws, col, row, header_row, data_dict, table_name)
            if data != '':
                column_array.append(data)
        paste_average(column_array, header_row, days_num, col, ws)
    ws.column_dimensions['A'].width = int(20)


def paste_data(ws, column, row, header_row, data_dict, table_name):
    """Using the current row for the date, and column to find teacher name
    it looks into the data_dict and matches."""
    curr_cell = ws.cell(row=row, column=column)
    date = ws.cell(row=row, column=1).value
    
    
    teacher_name = ws.cell(row=header_row, column=column).value
    teacher_name = teacher_name.replace('\r\n', ' ')
    date_dictionary = data_dict[date]
    teacher_data = date_dictionary[teacher_name]
    if "Day" in table_name:
        data = teacher_data['Day Average']
        curr_cell.value = data
    else:
        data = teacher_data['Night Average']
        curr_cell.value = data
    return data
    

def paste_average(column_array, header_row, days_num, col, ws):
    """Calculates the average and paste it in the bottom row of the 
    table"""
    if len(column_array) > 0:
        average = round(sum(column_array)/len(column_array), 2)
        avg_cell_row = header_row+days_num+1
        average_cell = ws.cell(row=avg_cell_row, column=col)
        big_font(average_cell, average)
