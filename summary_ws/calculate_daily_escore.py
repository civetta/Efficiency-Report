from daily_ws.calculate_block_escore import find_empty_row
from openpyxl.styles import Font


def find_non_empty_tables(wb,df):
    """Goes through each worksheet and finds each teacher table and
     the first empty row in each table"""
    week = wb.sheetnames
    week = week[:-1]   
    all_data = {}
    
    for day in week:
        day_dict = {}
        ws = wb[day]
        for col in range(8, ws.max_column+1):
            """So we go through the teachers in the ws, they start 8 columns 
            in, and use that number to calculate where to put the new table.
            So we do teacher location column -7, which should give us a
            1,2,3 count, and then times 8, so each teacher is located 8
            rows apart. Finally we subtract it by 6, so that the tables start
            higher up in the worksheet versus at row 8. So the first teacher is
            located at row 2, the second teacher is located at
            row 10, and so on"""
            table_row = ((col-7)*9)-6
            teacher = ws.cell(row=1, column=col).value
            empty_row = find_empty_row(ws, table_row)
            if table_row-empty_row != 0:
                #create_arrays(ws, table_row, empty_row, teacher, day_dict)
                paste_over(df, ws, empty_row, teacher, day_dict)
        all_data.update({day: day_dict})
        
    return all_data




def paste_over(df, ws, empty_row, teacher,day_dict):
    """Given the teacher and day it creates a new DF of just that information.
    Then it creates two sub DFS one where is_night is True and the other when it is 
    False. Then it calls paste_df_data to actually paste it into the excel file."""
    day = ws.cell(row=2, column=6).value
    day = day[:day.index(' ')].strip()
    df = df[(df.TeacherName == teacher)]
    df = df[(df.ws == ws.title)]
    day_df = df[(df.is_night == False)]
    night_df = df[(df.is_night == True)]
    bold = Font(bold=True)
    if not day_df.empty:
        day_avg = paste_df_data(empty_row,ws,day_df)
        ws.cell(row=empty_row, column=1, value="Day Total").font = bold
        ws.cell(row=empty_row, column=4, value=day_avg).font = bold
        empty_row=empty_row+1
    else:
        day_avg = ""
    if not night_df.empty:
        night_avg = paste_df_data(empty_row,ws,night_df)
        ws.cell(row=empty_row, column=1, value="Night Total").font = bold
        ws.cell(row=empty_row, column=4, value=night_avg).font = bold
    else:
        night_avg = ""
    return day_dict.update(
        {teacher: {'Day Average': day_avg, "Night Average": night_avg}})


def paste_df_data(empty_row, ws, df):       
    sum_students = df.Block.sum()
    sum_tabby = df.Tab.sum()
    if sum_tabby>0:
        day_average = round((sum_students/float(sum_tabby)),2)
        bold = Font(bold=True)
        ws.cell(row=empty_row, column=2, value = sum_students).font = bold
        ws.cell(row=empty_row, column=3, value = sum_tabby).font = bold
        ws.cell(row=empty_row, column=4, value = day_average).font = bold
        return day_average

