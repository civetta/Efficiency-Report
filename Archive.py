from openpyxl import Workbook
import os
import gspread
import oauth2client   

#Completes all of the authorization for google sheets
def authorize(name_of_gsheet):
    from oauth2client.service_account import ServiceAccountCredentials
    scope = "https://spreadsheets.google.com/feeds"
    credentials = ServiceAccountCredentials.from_json_keyfile_name('C:\\Users\\kheyden\\Documents\\Program\\archivekey.json', scope)
    gs = gspread.authorize(credentials)
    gsheet = gs.open("Test")
    wsheet = gsheet.worksheet(name_of_gsheet)
    return wsheet

#Creates excel worksheet
def creat_worksheets(wb):
    ws=wb.create_sheet('Raw Pulls')
    return ws
        
#Takes what is inside of the google spreadsheet and pastes it inside of the excel sheet   
def copy_into_excel(wsheet,ws,ranger,n):
    col_count = wsheet.col_count
    col = ranger[0]+1
    while col <= ranger[1]:
        excel_col=ws.max_column
        list_of_values = wsheet.col_values(col)
        list_of_values=convert_type(list_of_values)
        row_count = len(list_of_values)
        row = 1
        while row < row_count:     
            ws.cell(row = row, column = excel_col+n,value = list_of_values[row-1])
            row = row+1
        col = col+1
        n=1
        
#Converts all of the types from gsheet into correct types (string vs integer) so they are formatted in excel correctly.
def convert_type(list_of_values):
    for i in list_of_values:
            try:
                i=int(i)
            except:
                i=str(i)
    return list_of_values

#Find the first teacher and last teacher in the team, using team Leads as the splitters
def find_team_range(wsheet,start,end):
    names=wsheet.row_values(1)
    start=names.index(start)
    end=names.index(end)
    return [start,end]

#Calls all of the other functions above
def archive_to_excel(start,end,name_of_gsheet):
    wsheet = authorize(name_of_gsheet)
    wb = Workbook()
    ws=creat_worksheets(wb)
    ranger=find_team_range(wsheet,start,end)
    #Copies the date and tabby into excel
    copy_into_excel(wsheet,ws,[0,2],0)
    #Copies teachers data into excel
    copy_into_excel(wsheet,ws,ranger,1)
    return wb


    
