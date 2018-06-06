from datetime import datetime
from openpyxl.styles import PatternFill


def organize_data(ws, start, end, column, block, tab_list, max_col):
    """Declares all of the variables needed to create and organize the
    efficiency table"""
    average_student = round(sum(block)/float(len(block)), 2)
    average_tabby = round(sum(tab_list)/float(len(tab_list)), 2)
    block_escore = round(average_student/float(average_tabby), 2)
    teacher_name = str(ws.cell(row=1, column=column).value)
    start_time = str(ws.cell(row=start, column=1).value)
    end_time = str(ws.cell(row=end, column=1).value)
    time_range = create_time_range(start_time, end_time)
     
    paste_list = [time_range, average_student, average_tabby, block_escore]
    start_col = max_col+2
    paste_data(ws, paste_list, teacher_name, column, start_col)


def create_time_range(start, end):
    """
    Input: Two string of time. 
    Example: 03/21/18 Thu 8:27 AM and 03/21/18 Thu 9:48 AM.  
    Output: A string with just the hours (example: 8:27 AM - 9:48 AM").
    Times is marked with an asteriks if one of the times is after 8 AM
    but before 1AM. Or if the day of the week is Saturday or Sunday
    These mark the cells so that they are calculated differently then
    the "day time" teachers.
    """
    lister = [start, end]
    time_range = ""
    night_shift_start = datetime.strptime('8:00PM', "%I:%M%p")
    night_shift_end = datetime.strptime('2:00AM', "%I:%M%p")
    for item in lister:
        hour = item[13:]
        hour = datetime.strptime(hour, '%I:%M %p')
        date = item[9:12]
        is_weekend_teacher = date == 'Sat' or date == 'Sun'
        is_night_teacher = night_shift_start <= hour or hour <= night_shift_end
        if is_weekend_teacher or is_night_teacher:
            hour = str(hour)+'*'
        time_range = time_range + str(hour) + " - "
        
    return time_range[:-3]

      
def paste_data(ws, paste_list, teacher_name, column, start_col):
    """Paste the organized data under the daily summary tables.
    It also makes the cells purple if it has an asteriks(*)
    which indiciates a night or weekend shift."""  
    if teacher_name is not None:
        table_row = find_table(ws, teacher_name, start_col)
        empty_row = find_empty_row(ws, table_row, start_col)
        if empty_row is not None:
            if paste_list[0].find('*') > -1:
                paste_list[0].replace('*', '')
                hex_color = "c6c0ed"
            else:
                hex_color = "ffffff"
            for count in range(len(paste_list)):
                cell = ws.cell(row=empty_row, column=start_col+count)
                cell.value = paste_list[count]
                cell.fill = PatternFill("solid", fgColor=hex_color)


def find_table(ws, teacher_name, start_col):
    """Finds the location of the teacher's daily summary table in the column"""
    for col in range(1, start_col):
        if ws.cell(row=1, column=col).value == teacher_name:
            return col*8-20


def find_empty_row(ws, starting_row, column):
    """Finds the first empty row in the teacher's daily summary table 
    so it can paste the new data in it"""
    for row in range(starting_row, ws.max_row):
        if ws.cell(row=row, column=column).value is None:
            return row
        

def daily_array(teacher_name, max_col, ws):
    """Creates an Array of each blocks Efficiency Scores"""
    table_column = max_col+2
    table_row_start = find_table(ws, teacher_name, table_column)
    day_array = []
    night_array = []
    night_cell_color = PatternFill("solid", fgColor="c6c0ed")
    table_row_end = find_empty_row(ws, table_row_start, table_column)
    for i in range(table_row_start+2, table_row_end):
        if ws.cell(row=i, column=table_column).fill == night_cell_color:
            night_array.append(float(ws.cell(row=i, column=table_column+3).value))
        else:
            day_array.append(float(ws.cell(row=i, column=table_column+3).value))
print 

