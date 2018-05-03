import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def split_sheet_by_days(wb):
    search_day=["Mon","Tue","Wed","Thu","Fri","Sat","Sun","None","None2"]
    ws = wb.get_sheet_by_name("Raw Changes")
    max_row=ws.max_row
    max_column=ws.max_column
    day_count=0
    rower=1
    a=3
    a=find_start(ws,a,max_column,max_row)
    make_sheets(wb,ws,max_column,search_day[day_count])
    while a < max_row:
        while day_count<=6:
            try:
                day_sheet=wb.get_sheet_by_name(search_day[day_count])
            except:
                day_count=day_count+1
                break
            current_cell=ws.cell(row=a,column=1).value
            if ws.cell(row=a,column=1).value == None:
                break
            if search_day[day_count+1] and "6:00 AM" not in current_cell:
                rower=rower+1
                for col in range(1,max_column+1):
                    valv=ws.cell(row=a, column=col).value
                    day_sheet.cell(row=rower, column=col, value=valv)
            else:
                print "--------------------"
                day_count=day_count+1
                if search_day[day_count+1]=="None2":
                    break
                a=find_start(ws,a,max_column,max_row)
                if a == max_row:
                    break
                day=search_day[day_count]
                date_cell=str(ws.cell(row=a, column=1).value)
                date_cell=date_cell[date_cell.index(" "):].strip()
                date_cell=date_cell[:date_cell.index(" ")]
                date_cell=date_cell.strip()
                make_sheets(wb,ws,max_column,date_cell)
                rower=1
            a=a+1
    try:
        clean_up(wb)
    except:
        pass
    return wb


        
def find_start(ws,a,max_column,max_row):
    b=a
    while a<max_row:
        """date_cell=str(ws.cell(row=a, column=1).value)
        print date_cell
        if "12:55 AM" in date_cell:
            return b"""
        for col in range (3,max_column+1):
            val=ws.cell(row=a, column=col).value
            
            try:
                
                if int(val)>0:
                
                    val1=int(ws.cell(row=a+1, column=col).value)
                    val2=int(ws.cell(row=a+2, column=col).value)
                    val3=int(ws.cell(row=a+3, column=col).value)
                    val4=int(ws.cell(row=a+4, column=col).value)
                    
                    if val1>0 or val2>0 or val3>0 or val4>0:
                        print val1,val2,val3,val4
                        return a
            except:
                continue
        a=a+1
    return a

def clean_up(wb):
    search_day=["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    remove_list=[]
    for day in search_day:
            ws = wb.get_sheet_by_name(day)
            if ws.max_row ==1:
                remove_list.append(day)
    for empty in remove_list:
        ws = wb.get_sheet_by_name(empty)
        wb.remove_sheet(ws)
        
        
def make_sheets(wb,ws,max_column,day):
    day_sheet = wb.create_sheet(day)
    for a in range(max_column):
        val=ws.cell(row=1,column=a+1).value
        day_sheet.cell(row=1,column=a+1,value=val)

            
