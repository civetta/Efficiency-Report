from openpyxl import load_workbook


old = load_workbook(filename='Testing_Comparisons/12-10-18.xlsx')
new = load_workbook(filename='Testing_Comparisons/New-12-10-18.xlsx')

for sheet in old.sheetnames:
    row=1
    column=1
    old_sheet = old[sheet]
    new_sheet = new[sheet]
    while column <= old_sheet.max_column:
        row=1
        while row <= old_sheet.max_row:
            old_cell=old_sheet.cell(row=row, column=column).value
            new_cell=new_sheet.cell(row=row, column=column).value
            try:
                old_cell=int(old_cell)
                new_cell=int(new_cell)
            except:
                pass
            if old_cell != new_cell:
                
                print sheet
                print row, column
                print old_cell
                print new_cell
                print ""

            row=row+1
        column=column+1
