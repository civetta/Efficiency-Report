from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
def create_tables(ws):
        max_col = ws.max_column
        for column in range(8, max_col+1):
            table_start_cell = ((column-7)*8)-6
            teacher_name = ws.cell(row=1, column=column).value
            title_list = [teacher_name, "Students", "Tabby", "Efficiency Score"]
            for i in range(4):
                current_cell = ws.cell(row=table_start_cell, column=i+1)
                current_cell.value = title_list[i]
                current_cell.font = Font(bold=True)
                current_cell.fill = PatternFill("solid", fgColor='F2F2F2')
                

def format_sheet(ws):
    for i in range(1, ws.max_row):
        current_cell = ws.cell(row=i, column=5)
        current_cell.fill = PatternFill("solid", fgColor='F2F2F2')
    for col in range(1, ws.max_column):
        ws.column_dimensions[get_column_letter(col)].width = int(15)
    ws.column_dimensions['A'].width = int(30)
    ws.column_dimensions['F'].width = int(30)