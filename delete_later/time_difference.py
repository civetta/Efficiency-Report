
from openpyxl.utils import get_column_letter
"""This module takes the YTD numbers from Raw Pulls sheet and finds
 the differences every 6 minutes. It then puts those differences 
 in the Raw Changes worksheet"""


def first_two_columns(old_ws, new_ws, max_row):
    """Copies the first two columns which are the dates and tabby"""
    for row in range(1, max_row):
        for col in range(1, 3):
                val = old_ws.cell(row=row, column=col).value
                try:
                    val = int(val)
                except:
                    val = str(val)
                new_ws.cell(row=row, column=col, value=val)


def copy_teachernames(old_ws, new_ws):
    """Copies the first row which is just the names of the teachers"""
    for col in range(1, old_ws.max_column+1):
        new_ws.cell(row=1, column=col, value=old_ws.cell(row=1, column=col).value)


def find_difference(old_ws, new_ws, max_row, max_col):
    """Find the actual difference between the values. 
    Tells how many students a teacher closed every 6 minutes"""
    for col in range(3, max_col):
        cell1 = (old_ws.cell(row=2,column=col).value)
        for row in range(3, max_row):    
            cell2 = (old_ws.cell(row=row, column=col).value)
            try:
                difference = int(cell2)-int(cell1)
            except:
                difference = None
            new_ws.cell(row=row, column=col, value=difference)
            cell1 = cell2


def make_time_difference_sheet(wb):
    """Calls all of the above functions and formats the columns"""
    old_ws = wb.get_sheet_by_name("Raw Pulls")
    new_ws = wb.create_sheet('Raw Changes')    
    max_row = old_ws.max_row+1
    max_col = old_ws.max_column+1
    first_two_columns(old_ws, new_ws, max_row)
    copy_teachernames(old_ws, new_ws)
    find_difference(old_ws, new_ws, max_row, max_col)
    old_ws.column_dimensions[get_column_letter(1)].width = int(25)
    new_ws.column_dimensions[get_column_letter(1)].width = int(25)
    return wb
