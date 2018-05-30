import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



def split_sheet_by_days(wb,skip_days):
    """Find start of day, and end of day, and then copies everything in 
    between into a new sheet. If end of day returns none, it breaks"""
    raw_changes_ws = wb.get_sheet_by_name("Raw Changes")
    max_row = raw_changes_ws.max_row
    max_column = raw_changes_ws.max_column
    end_of_day_row = 0
    while True:
        current_row = end_of_day_row+1
        if current_row>max_row:
            break
        if skip_days in raw_changes_ws.cell(row=current_row,column=1).value:
            current_row=find_next_day(raw_changes_ws,current_row,max_row)
        current_row = find_start(raw_changes_ws, current_row, max_column, max_row)
        end_of_day_row = find_end(raw_changes_ws, current_row, max_column, max_row)
        if current_row == end_of_day_row:
            pass
        if end_of_day_row==None:
            end_of_day_row=max_row
        print raw_changes_ws.cell(row=end_of_day_row,column=1).value
        current_day=find_current_day(raw_changes_ws,current_row)
        make_sheets(wb,current_row,end_of_day_row,raw_changes_ws,max_column,current_day)
    return wb


def find_next_day(raw_change_ws, current_row,max_row):
    while current_row<max_row:
        current_day_time=raw_change_ws.cell(row=current_row,column=1).value
        if "12:54 AM" in current_day_time:
            return current_row
        else:
            current_row=current_row+1

def find_current_day(raw_changes_ws,current_row):
    """Splices down to day of week"""
    current_day_time=raw_changes_ws.cell(row=current_row,column=1).value
    current_day_time=current_day_time[current_day_time.index(" ")+1:]
    current_day=current_day_time[:current_day_time.index(" ")]
    return current_day


def find_end(raw_changes_ws,current_row,max_column,max_row):
    """Finds next instance of 12:45AM, used to define the end row, or end of day."""    
    while current_row<max_row:
        current_day_time=raw_changes_ws.cell(row=current_row,column=1).value
        if "12:54 AM" in current_day_time:
            return current_row
        else:
            current_row=current_row+1

       
def find_start(ws,a,max_column,max_row):
    """Find the first row with at least 3 consecutive values above 0"""
    b=a
    while a<max_row:
        for col in range (3,max_column+1):
            val1=ws.cell(row=a, column=col).value
            try:  
                if int(val1)>0:
                    val2=int(ws.cell(row=a+1, column=col).value)
                    val3=int(ws.cell(row=a+2, column=col).value)
                    if val2>0 or val3>0:
                        return a
            except:
                continue
        a=a+1
    return a


def make_sheets(wb,current_row,end_of_day_row,raw_changes_ws,max_column,day):
    """Creates Worksheet with Day of Week-Day of Month title syntax.
    Copies and Pastes from Raw Changes using Start Row and End Row as ranges."""
    name=raw_changes_ws.cell(row=current_row,column=1).value
    name=name.replace("/","-")
    current_sheet = wb.create_sheet(day+" "+name[:5])
    for column in range(1,max_column):
        current_sheet.cell(row=1,column=column,value=raw_changes_ws.cell(row=1,column=column).value)
        rower=2
        for row in range(current_row,end_of_day_row):
            current_sheet.cell(row=rower,column=column,value=raw_changes_ws.cell(row=row,column=column).value)
            rower=rower+1
    return current_sheet
