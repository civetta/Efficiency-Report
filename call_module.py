import openpyxl
from openpyxl import load_workbook
import numpy as np
import re
import os
from day_split import split_sheet_by_days
from blocker import define_blocks
from efficency_calculator import create_block_table
from break_down import create_summary
from weekly_summary import find_row
from time_difference import make_time_difference_sheet
from Archive import archive_to_excel
from formatter import formatter
from teacher_books import create_books



def create_dir(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)
    else:
        pass



def clean_up(wb):
    raw_sheet = wb.get_sheet_by_name("Raw Changes")
    first_date=str(raw_sheet.cell(row=2,column=1).value)
    date=first_date[:first_date.index(" ")]
    date=date.replace("/","-")

def save_files(wb,Lead_Name):
    date=get_date(wb)
    folder_location = os.path.join('C:\Users\kheyden\OneDrive - Imagine Learning\Reports\Efficiency Reports')
    create_dir(folder_location)
    lead_book_location=os.path.join(folder_location,"TEAM E-REPORTS")
    create_dir(lead_book_location)
    final_save_name = os.path.join(lead_book_location,date+"-TEAMWIDE.xlsx")
    wb.save(final_save_name)
    #create_books(max_col,wb,folder_location,date)
    

def get_date(wb):
    raw_sheet = wb.get_sheet_by_name("Raw Changes")
    first_date=str(raw_sheet.cell(row=2,column=1).value)
    date=first_date[:first_date.index(" ")]
    date=date.replace("/","-")
    return date


Lead_Names=["Jeremy Shock",'']
#"Jeremy Shock","Rachel Adams","Jairo  Rios","Salome Saenz","Kristin Donnelly","Caren Glowa",''
for x in range(len(Lead_Names)-1):
    print "------------------------------------------------------------"
    print Lead_Names[x]
    start_of_team=Lead_Names[x]
    end_of_team=Lead_Names[x+1]
    current_lead=Lead_Names[x]
    if current_lead=="Salome Saenz":
        current_lead="Jill Szafranski"
    name_of_gsheet="03/26"
    wb = archive_to_excel(start_of_team,end_of_team,name_of_gsheet)
    wb=make_time_difference_sheet(wb)
    wb=split_sheet_by_days(wb)
    wb=define_blocks(wb)
    max_col=create_block_table(wb,start_of_team)
    create_summary(wb,max_col)
    find_row(wb)
    std=wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(std)
    formatter(wb)
    clean_up(wb)
    folder_location=save_files(wb,current_lead)
