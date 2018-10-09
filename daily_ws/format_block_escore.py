from openpyxl.styles.borders import Border, Side
import openpyxl

def find_table(ws, teacher_name, max_col):
    """Finds the location of the teacher's daily summary table in the column"""
    for col in range(7, max_col+1):
        if ws.cell(row=1, column=col).value == teacher_name:
            return ((col-7)*8)-6


def find_empty_row(ws, starting_row):
    """Finds the first empty row in the teacher's daily summary table 
    so it can paste the new data in it"""
    if starting_row == ws.max_row:
        max_row = ws.max_row+6
    else:
        max_row = ws.max_row
    for row in range(starting_row, max_row):
        if starting_row==max_row:
            print "START = MAX"
        if ws.cell(row=row, column=1).value is None:
            return row


def coniditional_format_row(ws, empty_row, paste_list, scores, night_check):
    """Using the condition_list as scores, it conditionally formats
    "good scores' as green, 'bad scores' as red, and too good scores as
    blue"""
    if night_check is True:
        good_score = scores.get("Good Night Score")
    else:
        good_score = scores.get("Good Day Score")
    upper_bound = scores.get("Upper Bound")
    e_score = paste_list[-1]
    no_fill = openpyxl.styles.PatternFill(fill_type=None)
    return 'ffffff'
    #if e_score < good_score:
        #return 'f2b8ea'
    #elif e_score > upper_bound:
        #return 'c0f7f4'
    #else:
        #return 'dff7c0'


def night_time_teacher(ws, empty_row):
    """Creates thick borders around night time teacher cells"""
    thick_border = Border(left=Side(border_style='thick', color='1F49A1'),
                right=Side(border_style='thick', color='1F49A1'),
                top=Side(border_style='thick', color='1F49A1'),
                bottom=Side(border_style='thick', color='1F49A1'))
    for i in range(4):
        current_cell = ws.cell(row=empty_row, column=i+1)
        current_cell.border = thick_border
