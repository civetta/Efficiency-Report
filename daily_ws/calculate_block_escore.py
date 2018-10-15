from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from format_block_escore import find_empty_row, find_table
from format_block_escore import coniditional_format_row, night_time_teacher
import numpy as np


def organize_data(
        ws, start, end, column, block_list, tab_list, max_col, checks, scores,wb):
    """Declares all of the variables needed to create and organize the
    efficiency table. Then it calls paste_data to paste all of the data we have.
    This function also figures out if there is ever a night time shift and 
    return it back up to callmodule. This is used in the last module to know
    if we should create a day table, a night table, or both."""
    average_student = round(np.mean(block_list),2)
    average_tabby = round(np.mean(tab_list),2)
    block_escore = round(average_student/float(average_tabby), 2)
    teacher_name = str(ws.cell(row=1, column=column).value)
    start_time = str(ws.cell(row=start, column=6).value)
    end_time = str(ws.cell(row=end, column=6).value)
    time_range = create_time_range(start_time, end_time, ws)
    if '*' in time_range:
        checks['Night Check'] = True
    if '*' not in time_range:
        checks['Day Check'] = True
    paste_list = [time_range, average_student, average_tabby, block_escore]
    paste_data(ws, paste_list, teacher_name, max_col, scores,wb)
    return checks


def create_time_range(start, end, ws):
    """
    Input: Two string of time. 
    Example: 03/21/18 Thu 8:27 AM and 03/21/18 Thu 9:48 AM.  
    Output: A string with just the hours (example: 8:27 AM - 9:48 AM").
    Times is marked with an asteriks if one of the times is after 8 AM
    but before 1AM. Or if the day of the week is Saturday or Sunday
    These mark the cells so that they are calculated differently then
    the "day time" teachers.
    """
    start = datetime.strptime(start, '%m/%d/%y %a %I:%M %p')
    end = datetime.strptime(end, '%m/%d/%y %a %I:%M %p')
    nightshift_start = start.replace(hour=20, minute=0)
    nightshift_end = start + timedelta(days=1)     
    time_range = ""
    if end.weekday() >= 5:
        time_range = '*'
    elif end > nightshift_start and end < nightshift_end:
        time_range = '*'
    return time_range+start.strftime("%I:%M %p")+"-"+end.strftime("%I:%M %p")


def paste_data(ws, paste_list, teacher_name, max_col, scores,wb):
    """Paste the organized data under the daily summary tables.
    It also outlines the cell if it has an asteriks(*)
    which indiciates a night or weekend shift."""  
    night_check = False
    
    if teacher_name is not None:
        """Finds location of current active teachers, daily summary table"""

        starting_row = find_table(ws, teacher_name, max_col)
        empty_row = find_empty_row(ws, starting_row)
        """If there is an asteriks in the time_range (paste_list[0]) it is a 
        night shift and treated differently"""
        if '*' in paste_list[0]:
            night_time_teacher(ws, empty_row)
            night_check = True
        if paste_list[-1] is not None:
            color = coniditional_format_row(
                ws, empty_row, paste_list, scores, night_check)
        for i in range(4):
            """paste data from paste_list into empty row in daily summary table"""
            current_cell = ws.cell(row=empty_row, column=i+1)
            current_cell.value = paste_list[i]
            current_cell.fill = PatternFill("solid", fgColor=color)


