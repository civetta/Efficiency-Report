from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def call_create_tables(wb):
    """Goes through each day of the week worksheet and calls the functions
    below"""
    week = wb.get_sheet_names()
    #List of all of the worksheet names (which are the days of the week)
    week = week[:-2]
    for day in week:
        ws = wb.get_sheet_by_name(day)
        max_col = ws.max_column
        create_daily_tables(ws, max_col)
        format_sheet(ws)


def create_daily_tables(ws, max_col):
    """So we go through the teachers in the ws, they start 8 columns in, and
    use that number to calculate where to put the new table. So we do teacher
    location column -7, which should give us a 1,2,3 count, and then times 8,
    so each teacher is located 8 rows apart. Finally we subtract it by 6,
    so that the tables start higher up in the worksheet versus at row 8.
    So the first teacher is located at row 2, the second teacher is located at
    row 10, and so on"""
    for column in range(8, max_col+1):
        table_start_cell = ((column-7)*8)-6
        teacher_name = ws.cell(row=1, column=column).value
        title_list = [teacher_name, "Students", "Tabby", "Efficiency Score"]
        for i in range(4):
            current_cell = ws.cell(row=table_start_cell, column=i+1)
            current_cell.value = title_list[i]
            current_cell.font = Font(bold=True)
            current_cell.fill = PatternFill("solid", fgColor='F2F2F2')
            

def format_sheet(ws):
    """Goes through and grays out the spacer column. Makes the time stamp 
    columns extra wide so they are easy to read, and make all other columns
    at least 20 px in width"""
    for i in range(1, ws.max_row):
        current_cell = ws.cell(row=i, column=5)
        current_cell.fill = PatternFill("solid", fgColor='F2F2F2')
    for col in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width = int(20)
    ws.column_dimensions['A'].width = int(30)
    ws.column_dimensions['F'].width = int(30)