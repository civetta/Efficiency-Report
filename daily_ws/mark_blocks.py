from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from calculate_block_escore import organize_data
from datetime import datetime, timedelta
import pandas as pd
import numpy as np

def define_blocks(wb, checks, scores):
    """Goes through each ws and then goes through each column looking for the 
    start and end of each block. It then calls the bolder function
    conditionally format the cells, and then passes information over to 
    calculate_block_escore module to do calculations and paste data into 
    tables"""
    week = wb.get_sheet_names()
    week = week[:-1]
    all_df = pd.DataFrame(columns=['TeacherName','Block','Tab', 'TimeStamp'])
    new_rows={}
    for day in week:
        ws = wb.get_sheet_by_name(day)
        max_col = ws.max_column
        max_row = find_max_row(ws)
        col = 8
        start_row_to_look = 2
        #Each column is it's own teacher.
        while col <= max_col:
            teacher_name = ws.cell(row=1, column=col).value
            start = find_blocks(ws, col, max_row, start_row_to_look, 'start')
            if start != 'Next_Col' and start >= start_row_to_look and start != max_row :
                end = find_blocks(ws, col, max_row, start, 'end')
                start_row_to_look = end
                safe_to_color = empty_tabby(start,end,ws)
                if safe_to_color is True:
                    checks_and_lists = bolder(ws, start, end, col, max_col, checks, scores,wb)  
                    checks, block_df = checks_and_lists[0], checks_and_lists[1]
                    all_df = all_df.append(block_df)
            else:
                col = col+1
                start_row_to_look = 2
        live_metrics_down(ws,col,max_col,max_row)
    return [checks,all_df]


def live_metrics_down(ws,col,max_col,max_row):
    row=2
    for row in range(2,max_row):
        tab = ws.cell(row=row,column=7).value
        if tab == 0:
            for col in range(8,max_col):
                val = ws.cell(row=row,column=col).value
                if val > 0:
                    ws.cell(row=row,column=7,value="Metrics Down")
                    break


def find_max_row(ws):
    for row in range(1,ws.max_row):
        row_value = ws.cell(row=row,column=7).value
        if row_value == None:
            return row
    return ws.max_row


            
def empty_tabby(start,end,ws):
    tab_list = []
    for r in range(start, end):
        Tabby_Cell = ws.cell(row=r,  column=7).value
        if Tabby_Cell != None:
            Tabby_Cell = int(Tabby_Cell)
        tab_list.append(Tabby_Cell)
    average_tabby = round(sum(tab_list)/float(len(tab_list)), 2)
    if average_tabby>0:
        return True
    else:
        pass
        return False

def find_blocks(ws, col, max_row, starting_row, position):
    """Looks for either three 0's in a row for the end of a block,or 
    two sequential non zeros for the start of the block"""
    for row in range(starting_row, max_row):
        val1 = ws.cell(row=row, column=col).value
        if val1==None:
            return 'Next_Col'
        if position == 'start':  
            if int(val1) > 0:
                val2 = int(ws.cell(row=row+1, column=col).value)
                val3 = int(ws.cell(row=row+2, column=col).value)
                if val2 > 0 or val3 > 0:
                    return row
        elif position == 'end':
            if int(val1) == 0:
                try:
                    val2 = int(ws.cell(row=row+1, column=col).value)
                except TypeError:
                    break
                if val2 == 0:
                    return row
    return max_row


def bolder(ws, start, end, column, max_col, checks, scores,wb):
    """This function goes through and bolds and conditionally
    formatts each of the blocks. It then creates a list of 
    all of the information in the blocks. So Tab_list is a 
    list of all of the tabby's during the block, block_list is a list
    of all of the students taken during that block, and so on. This
    information is passed over into the calculate_block_escore module
    who pastes all of these data into daily teacher summary tables."""            
    tab_list = []
    block_list = []
    block_df = pd.DataFrame(columns=['TeacherName','Block','Tab', 'TimeStamp'])
    teacher_name = ws.cell(row=1, column=column).value
    for r in range(start, end):
        current_cell = ws.cell(row=r,  column=column)
        current_value = int(current_cell.value)
        Tabby_Cell = int(ws.cell(row=r,  column=7).value)
        plus_1_check = current_value == Tabby_Cell+1 
        minus_1_check = current_value == Tabby_Cell-1
        current_cell.font = Font(bold=True)
        if current_value == Tabby_Cell or plus_1_check or minus_1_check:
            current_cell.fill = PatternFill("solid", fgColor='dff7c0')
        elif current_value < Tabby_Cell and current_value > 0:
            current_cell.fill = PatternFill("solid", fgColor='f2b8ea')
        elif current_value >= Tabby_Cell+2:
            current_cell.fill = PatternFill("solid", fgColor='c0f7f4')
        else:
                pass
        if Tabby_Cell == 0:
            pass 
        else:
            time_cell = ws.cell(row=r, column=6).value
            row_in_block_df = pd.DataFrame({'TeacherName':[teacher_name],'Block':[current_value],'Tab':[Tabby_Cell],'TimeStamp': [time_cell]})
            block_df = block_df.append(row_in_block_df)
    checks = organize_data(
        ws, start, end, column, block_list, tab_list, max_col, checks, scores,wb)
    block_df = mark_as_night(block_df)
    return [checks, block_df]


def mark_as_night(block_df):
    """Formats block_df and checks to see if it's a night shift or not. This was added on 10/16/18, as a request 
    from management to use sum of all blocks/sum of all ss_max to calculate Efficiency Report. Instead of re-writing
    entire module I kept the old efficiency report and added in the df aspect in to work along side it"""
    block_df['TimeStamp'] = block_df['TimeStamp'].map(lambda x: datetime.strptime(x, "%m/%d/%y %a %I:%M %p"))
    block_df['Time'] = block_df['TimeStamp'].map(lambda x: x.strftime('%I:%M %p'))
    block_df['Date'] = block_df['TimeStamp'].map(lambda x: x.strftime('%m/%d/%y'))


    nightshift_start = block_df['TimeStamp'].iloc[0].replace(hour=20, minute=0)
    nightshift_end = block_df['TimeStamp'].iloc[0] + timedelta(days=1)
    end = block_df['TimeStamp'].iloc[-1]
    if end > nightshift_start and end < nightshift_end:
        block_df['is_night']=True
    else:
        block_df['is_night']=False
    return block_df