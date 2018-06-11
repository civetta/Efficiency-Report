from Time_Difference_03 import make_time_difference_sheet
from Mark_Blocks_05 import define_blocks
from Day_Split_04 import split_sheet_by_days
from openpyxl import load_workbook
import openpyxl
from Create_Daily_Escores import find_non_empty_tables
"""User Input Variables"""
skip_days = ['03/29', '3/27']
condition_list = {"Good Score": float(.90), "Upper Bound": float(1.25)}
print skip_days
"""Calling Functions"""
wb = load_workbook(filename='TestSource.xlsx')
print wb 
ws = wb.get_sheet_by_name('Raw Pulls')
print ws.cell(row=1,column=1).value
make_time_difference_sheet(wb)
print "Made Difference"
split_sheet_by_days(wb, skip_days)
print "Split Sheets"
define_blocks(wb)
print "Defined Blocks"
find_non_empty_tables(wb)
print "Completed Find Non Empty Tables"
blank_sheeet = wb.get_sheet_by_name('Sheet')
wb.remove_sheet(blank_sheeet)
wb.save('Test4.xlsx')